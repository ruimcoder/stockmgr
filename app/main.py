from __future__ import annotations

import io
import json
import logging
import math
import secrets
import zipfile
from contextlib import asynccontextmanager
from datetime import UTC, date, datetime, timedelta
from pathlib import Path
from typing import Any
from urllib.parse import quote as url_quote
from urllib.parse import urlencode, urlsplit, urlunsplit

import httpx
from authlib.integrations.base_client.errors import OAuthError
from authlib.integrations.starlette_client import OAuth
from authlib.jose.errors import JoseError
from fastapi import Depends, FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from jsonschema import ValidationError as JsonSchemaValidationError
from jsonschema import validate as jsonschema_validate
from pydantic import BaseModel, ValidationError
from sqlalchemy import func
from sqlmodel import Session, or_, select
from starlette.middleware.sessions import SessionMiddleware

from app.backup_utils import create_backup, list_backups, restore_backup
from app.config import get_settings
from app.db import _sqlite_path_from_url, get_session, init_db
from app.food_wheel import FOOD_GROUP_BY_KEY, FOOD_GROUPS, food_group_chart_data, infer_food_group
from app.i18n import SUPPORTED_LANGUAGES, translate
from app.models import LocationPlan, StockItem, StockMovement, User
from app.non_food_categories import ITEM_CATEGORIES, NON_FOOD_CATEGORIES
from app.uom_constants import UOM_OPTIONS
from app.pdf_utils import generate_table_pdf
from app.schemas import (
    BarcodeLookupRequest,
    BarcodeLookupResult,
    ExcelStockUpsertRequest,
    ExcelStockUpsertRow,
    ItemCreate,
    ItemRead,
    LocationPlanCreate,
)
from app.services.barcode import BarcodeLookupService
from app.services.calendar import CalendarSyncError, CalendarSyncService
from app.services.imports import parse_import_file
from app.version import APP_VERSION as _APP_VERSION, BUILD_DATE as _BUILD_DATE

settings = get_settings()
BASE_DIR = Path(__file__).resolve().parent
logger = logging.getLogger(__name__)

# Food group lookup map for templates (key → {name_en, name_pt, color})
_FOOD_GROUPS_MAP: dict[str, dict] = {
    fg.key: {"name_en": fg.name_en, "name_pt": fg.name_pt, "color": fg.color}
    for fg in FOOD_GROUPS
}

_DEFAULT_SECRET_KEY = "change-me-for-production"


@asynccontextmanager
async def lifespan(_: FastAPI):
    if settings.environment == "production":
        if settings.secret_key == _DEFAULT_SECRET_KEY:
            raise RuntimeError(
                "SECRET_KEY is set to the default insecure value. "
                "Set a strong SECRET_KEY environment variable before running in production."
            )
    db_file = _sqlite_path_from_url(settings.database_url)
    if db_file is not None:
        if not db_file.is_absolute():
            db_file = (Path.cwd() / db_file).resolve()
        create_backup(db_file)
    init_db()
    yield


app = FastAPI(title="stockmgr MVP", lifespan=lifespan)
app.add_middleware(
    SessionMiddleware,
    secret_key=settings.secret_key,
    same_site="lax",
    https_only=settings.environment == "production",
)
app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")

# Media uploads directory (product images)
_MEDIA_DIR = BASE_DIR.parent / "media"
_MEDIA_DIR.mkdir(parents=True, exist_ok=True)
app.mount("/media", StaticFiles(directory=str(_MEDIA_DIR)), name="media")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
templates.env.filters["urlencode_path"] = lambda v: url_quote(str(v), safe="")


def _datefmt(value: object) -> str:
    from datetime import date as _date, datetime as _datetime

    if value is None:
        return "-"
    if isinstance(value, (_date, _datetime)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, str) and value:
        try:
            return _date.fromisoformat(value).strftime("%d/%m/%Y")
        except ValueError:
            return value
    return str(value) if value else "-"


templates.env.filters["datefmt"] = _datefmt


def _datefmt(value: object) -> str:
    from datetime import date as _date, datetime as _datetime

    if value is None:
        return "-"
    if isinstance(value, (_date, _datetime)):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, str) and value:
        try:
            return _date.fromisoformat(value).strftime("%d/%m/%Y")
        except ValueError:
            return value
    return str(value) if value else "-"


templates.env.filters["datefmt"] = _datefmt

oauth = OAuth()
if settings.google_client_id and settings.google_client_secret:
    oauth.register(
        name="google",
        client_id=settings.google_client_id,
        client_secret=settings.google_client_secret,
        server_metadata_url="https://accounts.google.com/.well-known/openid-configuration",
        client_kwargs={
            "scope": (
                "openid email profile "
                "https://www.googleapis.com/auth/calendar.events "
                "https://www.googleapis.com/auth/calendar.readonly"
            )
        },
    )

if settings.microsoft_client_id and settings.microsoft_client_secret:
    oauth.register(
        name="microsoft",
        client_id=settings.microsoft_client_id,
        client_secret=settings.microsoft_client_secret,
        server_metadata_url=(
            "https://login.microsoftonline.com/common/v2.0/.well-known/openid-configuration"
        ),
        client_kwargs={
            "scope": "openid profile email offline_access User.Read Calendars.ReadWrite"
        },
    )

barcode_service = BarcodeLookupService(settings)
calendar_service = CalendarSyncService(settings)


def _current_language(request: Request) -> str:
    lang = request.session.get("lang", "en")
    if lang not in SUPPORTED_LANGUAGES:
        return "en"
    return lang


def _get_or_create_csrf_token(request: Request) -> str:
    """Return the CSRF token stored in the session, generating one if absent."""
    token = request.session.get("_csrf_token")
    if not token:
        token = secrets.token_hex(32)
        request.session["_csrf_token"] = token
    return token


async def _validate_csrf(request: Request) -> None:
    """Dependency: validate the CSRF token submitted with an HTML form POST."""
    session_token = request.session.get("_csrf_token")
    if not session_token:
        raise HTTPException(status_code=403, detail="CSRF protection: session token missing.")
    form = await request.form()
    form_token = str(form.get("_csrf_token", ""))
    if not secrets.compare_digest(session_token, form_token):
        raise HTTPException(status_code=403, detail="CSRF protection: token mismatch.")


def _render(
    request: Request,
    template_name: str,
    context: dict[str, Any] | None = None,
    status_code: int = 200,
):
    lang = _current_language(request)
    payload = {
        "request": request,
        "settings": settings,
        "lang": lang,
        "t": lambda key: translate(lang, key),
        "csrf_token": _get_or_create_csrf_token(request),
        "app_version_semantic": _APP_VERSION,
        "app_build_date": _BUILD_DATE,
        "food_groups_map": _FOOD_GROUPS_MAP,
        "non_food_categories": NON_FOOD_CATEGORIES,
        "item_categories": ITEM_CATEGORIES,
        "uom_options": UOM_OPTIONS,
    }
    if context:
        payload.update(context)
    return templates.TemplateResponse(request, template_name, payload, status_code=status_code)


def _admin_email_set() -> set[str]:
    values = [entry.strip().lower() for entry in settings.admin_emails.split(",")]
    return {entry for entry in values if entry}


def _count_approved_admins(session: Session) -> int:
    query = (
        select(func.count())
        .select_from(User)
        .where(
            User.approval_status == "approved",
            User.is_admin.is_(True),
        )
    )
    return int(session.exec(query).one() or 0)


def _new_user_status(email: str, session: Session) -> tuple[str, bool]:
    email_l = email.lower()
    admin_emails = _admin_email_set()
    if email_l in admin_emails:
        return "approved", True
    if _count_approved_admins(session) == 0:
        return "approved", True
    return "pending", False


def _current_user(request: Request, session: Session) -> User | None:
    user_id = request.session.get("user_id")
    if not user_id:
        return None
    return session.get(User, user_id)


def _require_user_or_redirect(request: Request, session: Session) -> User | RedirectResponse:
    user = _current_user(request, session)
    if not user:
        return RedirectResponse("/login", status_code=303)
    if user.approval_status != "approved":
        request.session.clear()
        return RedirectResponse("/login?m=account-not-approved", status_code=303)
    return user


def _require_api_user(request: Request, session: Session = Depends(get_session)) -> User:
    user = _current_user(request, session)
    if not user or user.approval_status != "approved":
        raise HTTPException(status_code=401, detail="Authentication required.")
    return user


def _require_admin_user(request: Request, session: Session = Depends(get_session)) -> User:
    user = _require_api_user(request, session)
    if not user.is_admin:
        raise HTTPException(status_code=403, detail="Admin access required.")
    return user


def _require_excel_api_user(request: Request, session: Session = Depends(get_session)) -> User:
    configured_key = settings.excel_api_key.strip()
    if not configured_key:
        raise HTTPException(status_code=503, detail="Excel API access is not configured.")

    provided_key = (
        request.headers.get("x-excel-api-key")
        or request.headers.get("x-api-key")
        or request.query_params.get("api_key")
    )
    if (provided_key or "").strip() != configured_key:
        raise HTTPException(status_code=401, detail="Invalid Excel API key.")

    target_email = (
        request.headers.get("x-excel-user-email")
        or request.headers.get("x-user-email")
        or request.query_params.get("user_email")
        or settings.excel_api_user_email
    )
    email = (target_email or "").strip().lower()
    if not email:
        raise HTTPException(
            status_code=400,
            detail="Excel API user email is required via header, query, or configuration.",
        )

    user = session.exec(select(User).where(User.email == email)).first()
    if not user or user.approval_status != "approved":
        raise HTTPException(status_code=404, detail="Excel API user not found or not approved.")
    return user


def _upsert_oauth_user(
    session: Session,
    *,
    provider: str,
    subject: str,
    email: str,
    display_name: str,
    access_token: str | None = None,
    refresh_token: str | None = None,
) -> User:
    query = select(User).where(User.oauth_provider == provider, User.oauth_subject == subject)
    existing = session.exec(query).first()
    now = datetime.now(UTC)
    if existing:
        existing.email = email
        existing.display_name = display_name
        if access_token:
            existing.access_token = access_token
        if refresh_token:
            existing.refresh_token = refresh_token
        session.add(existing)
        session.commit()
        session.refresh(existing)
        return existing

    approval_status, is_admin = _new_user_status(email, session)
    user = User(
        email=email,
        display_name=display_name,
        oauth_provider=provider,
        oauth_subject=subject,
        approval_status=approval_status,
        is_admin=is_admin,
        access_token=access_token,
        refresh_token=refresh_token,
        requested_at=now,
        approved_at=now if approval_status == "approved" else None,
    )
    session.add(user)
    session.commit()
    session.refresh(user)
    return user


def _parse_size(size_str: str | None) -> tuple[float | None, str | None]:
    """Parse a size string like '500g', '1.5kg', '250ml' into (value, unit)."""
    import re as _re
    if not size_str:
        return None, None
    m = _re.match(r"([0-9]+(?:[.,][0-9]+)?)\s*([a-zA-Z]+)", size_str.strip())
    if not m:
        return None, None
    try:
        val = float(m.group(1).replace(",", "."))
        unit = m.group(2).lower()
        return val, unit
    except ValueError:
        return None, None


def _item_payload_from_form(form: dict[str, Any]) -> dict[str, Any]:
    selected_storage_location = form.get("storage_location")
    if selected_storage_location == "__new__":
        form["storage_location"] = form.get("storage_location_new")
    payload: dict[str, Any] = {}
    for key in (
        "barcode",
        "batch_code",
        "name",
        "item_type",
        "storage_location",
        "storage_bucket",
        "expiry_date",
        "quantity",
        "unidose_per_pack",
        "target_unidoses_location",
        "temp_min_c",
        "temp_max_c",
        "humidity_min_pct",
        "humidity_max_pct",
        "renewal_date",
        "comment",
        "image_url",
        "nutriscore",
        "food_group",
        "weight_capacity",
        "uom",
        "item_category",
        "non_food_category",
    ):
        value = form.get(key)
        if key == "storage_bucket":
            payload[key] = value if value not in ("", None) else ""
        elif key == "quantity":
            payload[key] = int(value) if value not in ("", None) else 0
        elif key == "unidose_per_pack":
            payload[key] = int(value) if value not in ("", None) else 1
        elif key == "target_unidoses_location":
            payload[key] = int(value) if value not in ("", None) else 0
        elif key == "weight_capacity":
            payload[key] = float(value) if value not in ("", None) else None
        elif key in ("item_category", "non_food_category"):
            if value not in ("", None):
                payload[key] = value
        else:
            payload[key] = value if value not in ("", None) else None
    return payload


def _to_read_model(item: StockItem) -> ItemRead:
    return ItemRead.model_validate(item.model_dump())


def _excel_match_existing_item(session: Session, *, row: ExcelStockUpsertRow) -> StockItem | None:
    return session.exec(
        select(StockItem).where(
            StockItem.name == row.name,
            StockItem.item_type == row.item_type,
            StockItem.storage_location == row.storage_location,
            StockItem.storage_bucket == row.storage_bucket,
            StockItem.batch_code == row.batch_code,
            StockItem.expiry_date == row.expiry_date,
            StockItem.barcode == row.barcode,
        )
    ).first()


def _apply_item_create_to_stock(item: StockItem, item_in: ItemCreate) -> None:
    for key, value in item_in.model_dump().items():
        setattr(item, key, value)
    item.updated_at = datetime.now(UTC)


def _fetch_message(request: Request) -> str | None:
    return request.query_params.get("m")


def _looks_like_barcode(value: str) -> bool:
    normalized = value.replace(" ", "").strip()
    return normalized.isdigit() and 8 <= len(normalized) <= 20


def _product_batches(
    session: Session,
    *,
    item_type: str,
    product_name: str,
) -> list[StockItem]:
    batch_query = (
        select(StockItem)
        .where(
            StockItem.item_type == item_type,
            StockItem.name == product_name,
        )
        .order_by(StockItem.expiry_date, StockItem.storage_location, StockItem.storage_bucket)
    )
    return session.exec(batch_query).all()


def _storage_location_options(session: Session) -> list[str]:
    rows = session.exec(
        select(StockItem.storage_location)
        .where(StockItem.storage_location != "")
        .distinct()
        .order_by(func.lower(StockItem.storage_location), StockItem.storage_location)
    ).all()
    return [row for row in rows if row]


def _storage_location_field_context(
    session: Session, *, selected_location: str | None
) -> dict[str, Any]:
    location_options = _storage_location_options(session)
    selected = (selected_location or "").strip()
    if selected and selected in location_options:
        return {
            "storage_location_options": location_options,
            "storage_location_selection": selected,
            "storage_location_new_value": "",
        }
    if selected:
        return {
            "storage_location_options": location_options,
            "storage_location_selection": "__new__",
            "storage_location_new_value": selected,
        }
    return {
        "storage_location_options": location_options,
        "storage_location_selection": "",
        "storage_location_new_value": "",
    }


def _plan_locations(session: Session) -> list[str]:
    """Return sorted unique location names from LocationPlans + existing StockItems."""
    plan_locs = session.exec(
        select(LocationPlan.location).distinct().order_by(LocationPlan.location)
    ).all()
    item_locs = session.exec(
        select(StockItem.storage_location)
        .where(StockItem.storage_location != "")
        .distinct()
        .order_by(func.lower(StockItem.storage_location))
        .order_by(func.lower(StockItem.storage_location))
    ).all()
    seen: set[str] = set()
    result: list[str] = []
    for loc in list(plan_locs) + [r for r in item_locs if r]:
        key = (loc or "").strip()
        if key and key not in seen:
            seen.add(key)
            result.append(key)
    return result


def _lookup_provider_names() -> list[str]:
    """Return display names of all enabled barcode providers, in chain order."""
    cfg = barcode_service.config
    providers = cfg.get("providers", {})
    seen: set[str] = set()
    names: list[str] = []
    for chain in cfg.get("lookup", {}).get("chains", {}).values():
        for pid in chain:
            if pid not in seen:
                seen.add(pid)
                p = providers.get(pid, {})
                if p.get("enabled") and p.get("displayName"):
                    names.append(p["displayName"])
    return names


def _upsert_location_plan(session: Session, *, location: str, participants: int, stock_duration_days: int) -> None:
    """Create or update a LocationPlan for the given location."""
    location = location.strip()
    if not location:
        return
    existing = session.exec(select(LocationPlan).where(LocationPlan.location == location)).first()
    if existing:
        existing.participants = participants
        existing.stock_duration_days = stock_duration_days
        existing.updated_at = datetime.now(UTC)
        session.add(existing)
    else:
        session.add(LocationPlan(location=location, participants=participants, stock_duration_days=stock_duration_days))
    session.commit()


def _telegram_operation_message(*, operation: str, actor: str, detail: str) -> str:
    return f"[stockmgr] {operation}\nactor: {actor}\n{detail}"


def _notify_telegram_operation_sync(*, operation: str, actor: str, detail: str) -> None:
    logger.info(
        "Telegram app-level notifications are disabled; operation=%s actor=%s detail=%s",
        operation,
        actor,
        detail,
    )


async def _notify_telegram_operation_async(*, operation: str, actor: str, detail: str) -> None:
    logger.info(
        "Telegram app-level notifications are disabled; operation=%s actor=%s detail=%s",
        operation,
        actor,
        detail,
    )


def _telegram_help_message() -> str:
    return (
        "stockmgr Telegram commands:\n"
        "/help - command list\n"
        "/health - app health and deployed version\n"
        "/inventory - shared inventory summary\n"
        "/find <name> - search products by name\n"
        "/moves [N] - latest stock movements (default 5)\n"
        "You can also send plain text like: inventory, health, moves 10, find rice, or just 'rice'."
    )


def _telegram_inventory_summary(session: Session) -> str:
    batch_count = int(session.exec(select(func.count()).select_from(StockItem)).one() or 0)
    product_count = int(session.exec(select(func.count(func.distinct(StockItem.name)))).one() or 0)
    total_quantity = int(
        session.exec(select(func.coalesce(func.sum(StockItem.quantity), 0))).one() or 0
    )
    today = date.today()
    renewal_cutoff = today + timedelta(days=settings.renewal_window_days)
    renewal_count = int(
        session.exec(
            select(func.count())
            .select_from(StockItem)
            .where(
                StockItem.renewal_date.is_not(None),
                StockItem.renewal_date >= today,
                StockItem.renewal_date <= renewal_cutoff,
            )
        ).one()
        or 0
    )
    return (
        "Shared inventory summary:\n"
        f"Products: {product_count}\n"
        f"Batches: {batch_count}\n"
        f"Total quantity: {total_quantity}\n"
        f"Renewals in next {settings.renewal_window_days} days: {renewal_count}"
    )


def _telegram_find_products(session: Session, query: str) -> str:
    term = query.strip().lower()
    if not term:
        return "Usage: /find <product name>"
    rows = session.exec(
        select(
            StockItem.name,
            StockItem.item_type,
            func.sum(StockItem.quantity).label("total_quantity"),
        )
        .where(func.lower(StockItem.name).like(f"%{term}%"))
        .group_by(StockItem.name, StockItem.item_type)
        .order_by(func.sum(StockItem.quantity).desc(), StockItem.name)
        .limit(5)
    ).all()
    if not rows:
        return f"No products matched '{query.strip()}'."
    lines = ["Top matches:"]
    for name, item_type, total_quantity in rows:
        lines.append(f"- {name} ({item_type}) qty={int(total_quantity or 0)}")
    return "\n".join(lines)


def _telegram_recent_moves(session: Session, requested_limit: str) -> str:
    limit = 5
    normalized = requested_limit.strip()
    if normalized:
        if not normalized.isdigit():
            return "Usage: /moves [N]"
        limit = max(1, min(int(normalized), 20))
    rows = session.exec(
        select(StockMovement, StockItem)
        .join(StockItem, StockItem.id == StockMovement.stock_item_id)
        .order_by(StockMovement.created_at.desc())
        .limit(limit)
    ).all()
    if not rows:
        return "No stock movements recorded yet."
    lines = ["Latest stock movements:"]
    for movement, item in rows:
        sign = "+" if movement.delta >= 0 else ""
        note = f" ({movement.note})" if movement.note else ""
        lines.append(
            f"- {movement.created_at.date()} {item.name} {sign}{movement.delta}{note}"
        )
    return "\n".join(lines)


def _parse_telegram_input(text: str) -> tuple[str, str]:
    normalized = " ".join(text.strip().split())
    if not normalized:
        return "", ""

    if normalized.startswith("/"):
        parts = normalized.split(maxsplit=1)
        command = parts[0].lower().split("@", 1)[0]
        argument = parts[1] if len(parts) > 1 else ""
        return command, argument

    lowered = normalized.lower()
    if lowered in {"help", "commands", "menu", "start"}:
        return "/help", ""
    if lowered in {"health", "status", "ping"}:
        return "/health", ""
    if lowered in {"inventory", "stock", "inventory summary", "stock summary", "summary"}:
        return "/inventory", ""
    if lowered.startswith("find "):
        return "/find", normalized.split(maxsplit=1)[1]
    if lowered.startswith("search "):
        return "/find", normalized.split(maxsplit=1)[1]
    if lowered.startswith("lookup "):
        return "/find", normalized.split(maxsplit=1)[1]
    if lowered in {"moves", "recent moves", "latest moves", "movements"}:
        return "/moves", ""
    if lowered.startswith("moves "):
        return "/moves", normalized.split(maxsplit=1)[1]
    if lowered.startswith("recent moves "):
        return "/moves", normalized.split(maxsplit=2)[2]
    if lowered.startswith("latest moves "):
        return "/moves", normalized.split(maxsplit=2)[2]

    return "/find", normalized


def _handle_telegram_command(session: Session, text: str) -> str:
    command, argument = _parse_telegram_input(text)
    if not command:
        return "Unknown input. Send /help to list supported commands."

    if command in {"/start", "/help"}:
        return _telegram_help_message()
    if command == "/health":
        return f"Health: ok\nVersion: {settings.app_version}"
    if command == "/inventory":
        return _telegram_inventory_summary(session)
    if command == "/find":
        return _telegram_find_products(session, argument)
    if command == "/moves":
        return _telegram_recent_moves(session, argument)

    return "Unsupported command. Send /help to list supported commands."


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "version": settings.app_version}


@app.get("/manifest.webmanifest", include_in_schema=False)
def manifest() -> FileResponse:
    return FileResponse(
        BASE_DIR / "static" / "manifest.webmanifest",
        media_type="application/manifest+json",
    )


@app.get("/service-worker.js", include_in_schema=False)
def service_worker() -> FileResponse:
    return FileResponse(
        BASE_DIR / "static" / "service-worker.js",
        media_type="application/javascript",
    )


@app.get("/offline.html", include_in_schema=False)
def offline_page() -> FileResponse:
    return FileResponse(
        BASE_DIR / "static" / "offline.html",
        media_type="text/html",
    )


@app.get("/lang/{lang_code}")
def set_language(lang_code: str, request: Request):
    if lang_code not in SUPPORTED_LANGUAGES:
        raise HTTPException(status_code=404, detail="Unsupported language.")
    request.session["lang"] = lang_code
    return RedirectResponse(request.headers.get("referer") or "/", status_code=303)


@app.get("/register")
def register_page(request: Request, session: Session = Depends(get_session)):
    user = _current_user(request, session)
    if user and user.approval_status == "approved":
        return RedirectResponse("/", status_code=303)
    return _render(request, "register.html", {"message": _fetch_message(request)})


@app.post("/register")
def register_user(
    request: Request,
    email: str = Form(...),
    display_name: str = Form(...),
    session: Session = Depends(get_session),
):
    email_l = email.strip().lower()
    query = select(User).where(User.oauth_provider == "dev", User.oauth_subject == email_l)
    existing = session.exec(query).first()
    if existing:
        return RedirectResponse("/login?m=register-already-exists", status_code=303)

    now = datetime.now(UTC)
    status, is_admin = _new_user_status(email_l, session)
    user = User(
        email=email_l,
        display_name=display_name.strip() or email_l,
        oauth_provider="dev",
        oauth_subject=email_l,
        approval_status=status,
        is_admin=is_admin,
        requested_at=now,
        approved_at=now if status == "approved" else None,
    )
    session.add(user)
    session.commit()
    session.refresh(user)

    if user.approval_status == "approved":
        request.session["user_id"] = user.id
        return RedirectResponse("/?m=registered-approved", status_code=303)
    return RedirectResponse("/login?m=registration-pending", status_code=303)


@app.get("/login")
def login_page(request: Request, session: Session = Depends(get_session)):
    user = _current_user(request, session)
    if user and user.approval_status == "approved":
        return RedirectResponse("/", status_code=303)
    available_oauth = [name for name in ("google", "microsoft") if oauth.create_client(name)]
    return _render(
        request,
        "login.html",
        {
            "available_oauth": available_oauth,
            "auth_mode": settings.auth_mode,
            "message": _fetch_message(request),
        },
    )


@app.post("/auth/dev-login")
def dev_login(
    request: Request,
    email: str = Form(...),
    session: Session = Depends(get_session),
):
    if settings.auth_mode != "dev":
        raise HTTPException(status_code=403, detail="Development login is disabled.")

    email_l = email.strip().lower()
    query = select(User).where(User.oauth_provider == "dev", User.oauth_subject == email_l)
    user = session.exec(query).first()
    if not user:
        return RedirectResponse("/register?m=login-register-first", status_code=303)
    if user.approval_status != "approved":
        return RedirectResponse("/login?m=login-pending-approval", status_code=303)
    request.session["user_id"] = user.id
    return RedirectResponse("/", status_code=303)


@app.get("/auth/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)


@app.get("/auth/{provider}/start")
async def oauth_start(request: Request, provider: str):
    client = oauth.create_client(provider)
    if not client:
        raise HTTPException(status_code=404, detail=f"OAuth provider '{provider}' is unavailable.")
    redirect_uri = _oauth_callback_redirect_uri(request, provider)
    return await client.authorize_redirect(request, redirect_uri)


def _oauth_callback_redirect_uri(request: Request, provider: str) -> str:
    callback_uri = str(request.url_for("oauth_callback", provider=provider))
    return _build_oauth_redirect_uri(
        callback_uri=callback_uri,
        public_base_url=settings.public_base_url,
        forwarded_proto=request.headers.get("x-forwarded-proto"),
    )


def _build_oauth_redirect_uri(
    callback_uri: str,
    public_base_url: str | None,
    forwarded_proto: str | None,
) -> str:
    parsed_callback = urlsplit(callback_uri)
    configured_base = (public_base_url or "").strip()
    if configured_base:
        if "://" not in configured_base:
            configured_base = f"https://{configured_base}"
        parsed_base = urlsplit(configured_base)
        if parsed_base.scheme and parsed_base.netloc:
            base_path = parsed_base.path.rstrip("/")
            callback_path = parsed_callback.path
            combined_path = f"{base_path}{callback_path}" if base_path else callback_path
            return urlunsplit(
                (
                    parsed_base.scheme,
                    parsed_base.netloc,
                    combined_path,
                    parsed_callback.query,
                    parsed_callback.fragment,
                )
            )

    normalized_forwarded_proto = (forwarded_proto or "").split(",")[0].strip().lower()
    if normalized_forwarded_proto == "https" and parsed_callback.scheme != "https":
        return urlunsplit(
            (
                "https",
                parsed_callback.netloc,
                parsed_callback.path,
                parsed_callback.query,
                parsed_callback.fragment,
            )
        )
    return callback_uri


def _oauth_message_from_error(error: str | None) -> str:
    normalized_error = (error or "").strip().lower()
    if normalized_error in {"access_denied", "user_cancelled"}:
        return "oauth-cancelled"
    if normalized_error:
        return "oauth-provider-error"
    return "oauth-login-failed"


@app.get("/auth/{provider}/callback")
async def oauth_callback(
    request: Request,
    provider: str,
    session: Session = Depends(get_session),
):
    client = oauth.create_client(provider)
    if not client:
        raise HTTPException(status_code=404, detail=f"OAuth provider '{provider}' is unavailable.")

    oauth_error = request.query_params.get("error")
    if oauth_error:
        message_key = _oauth_message_from_error(oauth_error)
        return RedirectResponse(f"/login?m={message_key}", status_code=303)

    token_kwargs: dict[str, Any] = {}
    if provider == "microsoft":
        # Microsoft /common may return tenant-specific issuers; accept token claims
        # without strict issuer matching and rely on subsequent Graph profile lookup.
        token_kwargs["claims_options"] = {}
    try:
        token = await client.authorize_access_token(request, **token_kwargs)
    except OAuthError as exc:
        logger.warning(
            "OAuth token exchange failed for provider %s: %s (%s)",
            provider,
            exc.error,
            exc.description,
        )
        message_key = _oauth_message_from_error(exc.error)
        return RedirectResponse(f"/login?m={message_key}", status_code=303)
    except JoseError as exc:
        logger.warning("OAuth ID token validation failed for provider %s: %s", provider, exc)
        return RedirectResponse("/login?m=oauth-provider-error", status_code=303)
    userinfo = token.get("userinfo")
    if not userinfo:
        if provider == "google":
            try:
                userinfo = await client.userinfo(token=token)
            except OAuthError as exc:
                logger.warning(
                    "OAuth userinfo lookup failed for provider %s: %s (%s)",
                    provider,
                    exc.error,
                    exc.description,
                )
                message_key = _oauth_message_from_error(exc.error)
                return RedirectResponse(f"/login?m={message_key}", status_code=303)
        elif provider == "microsoft":
            try:
                async with httpx.AsyncClient(timeout=8) as http_client:
                    graph_response = await http_client.get(
                        "https://graph.microsoft.com/v1.0/me",
                        headers={"Authorization": f"Bearer {token.get('access_token')}"},
                    )
                graph_response.raise_for_status()
                userinfo = graph_response.json()
            except httpx.HTTPError as exc:
                logger.warning("Microsoft Graph profile lookup failed: %s", exc)
                return RedirectResponse("/login?m=oauth-provider-error", status_code=303)
        else:
            raise HTTPException(status_code=400, detail="Unsupported OAuth provider.")

    email = userinfo.get("email") or userinfo.get("mail") or userinfo.get("userPrincipalName")
    if not email:
        raise HTTPException(
            status_code=400, detail="OAuth profile does not include an email address."
        )
    subject = str(userinfo.get("sub") or userinfo.get("id") or email)
    display_name = userinfo.get("name") or email.split("@")[0]

    user = _upsert_oauth_user(
        session,
        provider=provider,
        subject=subject,
        email=email.lower(),
        display_name=display_name,
        access_token=token.get("access_token"),
        refresh_token=token.get("refresh_token"),
    )
    if user.approval_status != "approved":
        return RedirectResponse("/login?m=login-pending-approval", status_code=303)
    request.session["user_id"] = user.id
    return RedirectResponse("/", status_code=303)


@app.get("/")
def index(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    bucket_filter = request.query_params.get("bucket_filter", "all")
    location_filter = request.query_params.get("location_filter", "").strip()
    statement = select(StockItem)
    if bucket_filter == "assigned":
        statement = statement.where(StockItem.storage_bucket != "")
    elif bucket_filter == "unassigned":
        statement = statement.where(StockItem.storage_bucket == "")
    if location_filter:
        statement = statement.where(StockItem.storage_location == location_filter)
    statement = statement.order_by(StockItem.name, StockItem.batch_code, StockItem.expiry_date)
    items = session.exec(statement).all()
    locations = session.exec(
        select(StockItem.storage_location)
        .distinct()
        .order_by(StockItem.storage_location)
    ).all()
    return _render(
        request,
        "index.html",
        {
            "user": user,
            "items": items,
            "locations": locations,
            "bucket_filter": bucket_filter,
            "location_filter": location_filter,
            "message": _fetch_message(request),
        },
    )


@app.post("/items/search")
def search_item_from_home(
    request: Request,
    query: str = Form(...),
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    search_value = query.strip()
    if not search_value:
        return RedirectResponse("/?m=search-empty", status_code=303)

    is_barcode = _looks_like_barcode(search_value)
    barcode_value = search_value.replace(" ", "").strip() if is_barcode else ""

    if is_barcode:
        barcode_match = session.exec(
            select(StockItem)
            .where(
                StockItem.barcode == barcode_value,
                StockItem.quantity > 0,
            )
            .order_by(StockItem.expiry_date, StockItem.updated_at.desc())
        ).first()
        if barcode_match:
            detail_url = (
                str(request.base_url).rstrip("/")
                + "/products/by-name/"
                + url_quote(barcode_match.item_type, safe="")
                + "/"
                + url_quote(barcode_match.name, safe="")
            )
            return RedirectResponse(detail_url, status_code=303)

        new_item_url = f"{request.url_for('item_new')}?{urlencode({'barcode': barcode_value})}"
        return RedirectResponse(new_item_url, status_code=303)

    name_match = session.exec(
        select(StockItem)
        .where(
            func.lower(StockItem.name) == search_value.lower(),
            StockItem.quantity > 0,
        )
        .order_by(StockItem.expiry_date, StockItem.updated_at.desc())
    ).first()
    if not name_match:
        name_match = session.exec(
            select(StockItem)
            .where(
                StockItem.name.like(f"%{search_value}%"),
                StockItem.quantity > 0,
            )
            .order_by(StockItem.name, StockItem.expiry_date, StockItem.updated_at.desc())
        ).first()

    if name_match:
        detail_url = (
            str(request.base_url).rstrip("/")
            + "/products/by-name/"
            + url_quote(name_match.item_type, safe="")
            + "/"
            + url_quote(name_match.name, safe="")
        )
        return RedirectResponse(detail_url, status_code=303)

    new_item_url = f"{request.url_for('item_new')}?{urlencode({'name': search_value})}"
    return RedirectResponse(new_item_url, status_code=303)


@app.get("/stock/views")
def stock_views(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user

    plans: dict[str, LocationPlan] = {
        p.location: p for p in session.exec(select(LocationPlan)).all()
    }

    overall_query = (
        select(
            StockItem.name,
            StockItem.item_type,
            func.sum(StockItem.quantity).label("quantity"),
            func.max(StockItem.nutriscore).label("nutriscore"),
            func.max(StockItem.food_group).label("food_group"),
        )
        .group_by(StockItem.name, StockItem.item_type)
        .order_by(StockItem.name)
    )
    by_location_query = (
        select(
            StockItem.name,
            StockItem.item_type,
            StockItem.storage_location,
            StockItem.food_group,
            func.sum(StockItem.quantity).label("quantity"),
            func.sum(StockItem.quantity * StockItem.unidose_per_pack).label("total_unidoses"),
            func.max(StockItem.target_unidoses_location).label("target_unidoses"),
            func.max(StockItem.nutriscore).label("nutriscore"),
        )
        .group_by(StockItem.name, StockItem.item_type, StockItem.storage_location)
        .order_by(StockItem.name, StockItem.storage_location)
    )
    by_location_expiry_query = (
        select(
            StockItem.name,
            StockItem.item_type,
            StockItem.storage_location,
            StockItem.expiry_date,
            func.sum(StockItem.quantity).label("quantity"),
            func.sum(StockItem.quantity * StockItem.unidose_per_pack).label("total_unidoses"),
            func.max(StockItem.nutriscore).label("nutriscore"),
            func.max(StockItem.food_group).label("food_group"),
        )
        .group_by(
            StockItem.name,
            StockItem.item_type,
            StockItem.storage_location,
            StockItem.expiry_date,
        )
        .order_by(StockItem.name, StockItem.storage_location, StockItem.expiry_date)
    )

    # Enrich location_rows with plan-based target
    raw_location = session.exec(by_location_query).all()
    location_rows = []
    for row in raw_location:
        name, item_type, location, food_group, qty, total_u, target_u, nutriscore = row
        plan = plans.get(location)
        if plan and food_group:
            fg = FOOD_GROUP_BY_KEY.get(food_group)
            plan_target = (
                math.ceil(plan.total_meal_occasions * fg.target_pct / 100) if fg
                else plan.total_meal_occasions
            )
        elif plan:
            plan_target = plan.total_meal_occasions
        else:
            plan_target = None
        effective_target = plan_target if plan_target is not None else int(target_u or 0)
        location_rows.append({
            "name": name,
            "item_type": item_type,
            "location": location,
            "quantity": qty,
            "total_unidoses": int(total_u or 0),
            "target_unidoses": effective_target,
            "plan_target": plan_target,
            "delta_unidoses": (effective_target - int(total_u or 0)) if effective_target else None,
            "nutriscore": nutriscore,
        })

    return _render(
        request,
        "stock_views.html",
        {
            "user": user,
            "overall_rows": session.exec(overall_query).all(),
            "location_rows": location_rows,
            "validity_rows": session.exec(by_location_expiry_query).all(),
        },
    )


@app.get("/shopping-list")
def shopping_list(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user

    # Load location plans for plan-based target computation
    plans: dict[str, LocationPlan] = {
        p.location: p for p in session.exec(select(LocationPlan)).all()
    }

    location_rows = session.exec(
        select(
            StockItem.name,
            StockItem.item_type,
            StockItem.storage_location,
            StockItem.food_group,
            func.sum(StockItem.quantity * StockItem.unidose_per_pack).label("total_unidoses"),
            func.max(StockItem.target_unidoses_location).label("target_unidoses"),
            func.max(StockItem.unidose_per_pack).label("unidose_per_pack"),
        )
        .group_by(StockItem.name, StockItem.item_type, StockItem.storage_location)
        .order_by(StockItem.name, StockItem.storage_location)
    ).all()

    grouped: dict[tuple[str, str], dict[str, Any]] = {}
    for row in location_rows:
        name, item_type, location, food_group, total_u, target_u, per_pack = row
        total_unidoses = int(total_u or 0)
        per_pack_value = max(1, int(per_pack or 1))

        # Prefer plan-based target if a plan exists for this location
        plan = plans.get(location)
        if plan and food_group:
            fg = FOOD_GROUP_BY_KEY.get(food_group)
            effective_target = (
                math.ceil(plan.total_meal_occasions * fg.target_pct / 100) if fg
                else plan.total_meal_occasions
            )
        else:
            effective_target = int(target_u or 0)

        delta_unidoses = max(effective_target - total_unidoses, 0)
        qty_to_buy = math.ceil(delta_unidoses / per_pack_value) if delta_unidoses else 0
        key = (name, item_type)
        if key not in grouped:
            fg_obj = FOOD_GROUP_BY_KEY.get(food_group) if food_group else None
            grouped[key] = {
                "name": name,
                "item_type": item_type,
                "food_group": food_group,
                "food_group_color": fg_obj.color if fg_obj else None,
                "total_quantity_to_buy": 0,
                "distribution": [],
            }
        grouped[key]["total_quantity_to_buy"] += qty_to_buy
        if qty_to_buy > 0:
            grouped[key]["distribution"].append(f"{location}: {qty_to_buy}")

    rows = [value for value in grouped.values() if value["total_quantity_to_buy"] > 0]
    rows.sort(key=lambda item: (item["name"], item["item_type"]))

    # Collect one image_url per product for hover cards
    image_rows = session.exec(
        select(StockItem.name, StockItem.item_type, StockItem.image_url)
        .where(StockItem.image_url.isnot(None))
        .distinct()
    ).all()
    image_lookup: dict[tuple[str, str], str] = {}
    for r_name, r_type, r_img in image_rows:
        key = (r_name, r_type)
        if key not in image_lookup and r_img:
            image_lookup[key] = r_img
    for row in rows:
        row["image_url"] = image_lookup.get((row["name"], row["item_type"]))

    return _render(request, "shopping_list.html", {"user": user, "rows": rows})


# ── Location plans ──────────────────────────────────────────────────────────


@app.get("/location-plans")
def location_plans_list(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    plans = session.exec(select(LocationPlan).order_by(LocationPlan.location)).all()
    return _render(
        request,
        "location_plans.html",
        {
            "user": maybe_user,
            "plans": plans,
            "edit_plan": None,
            "message": _fetch_message(request),
        },
    )


@app.post("/location-plans")
async def location_plans_create(
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    form = await request.form()
    try:
        plan_in = LocationPlanCreate(
            location=str(form.get("location") or ""),
            participants=int(form.get("participants") or 1),
            stock_duration_days=int(form.get("stock_duration_days") or 1),
        )
    except (ValueError, Exception) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    existing = session.exec(
        select(LocationPlan).where(LocationPlan.location == plan_in.location)
    ).first()
    if existing:
        existing.participants = plan_in.participants
        existing.stock_duration_days = plan_in.stock_duration_days
        existing.updated_at = datetime.now(UTC)
        session.add(existing)
    else:
        session.add(LocationPlan(**plan_in.model_dump()))
    session.commit()
    return RedirectResponse("/location-plans?m=location-plan-saved", status_code=303)


@app.get("/location-plans/{plan_id}/edit")
def location_plans_edit(
    plan_id: int, request: Request, session: Session = Depends(get_session)
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    plan = session.get(LocationPlan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Location plan not found.")
    all_plans = session.exec(select(LocationPlan).order_by(LocationPlan.location)).all()
    return _render(
        request,
        "location_plans.html",
        {"user": maybe_user, "plans": all_plans, "edit_plan": plan, "message": None},
    )


@app.post("/location-plans/{plan_id}/update")
async def location_plans_update(
    plan_id: int,
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    plan = session.get(LocationPlan, plan_id)
    if not plan:
        raise HTTPException(status_code=404, detail="Location plan not found.")
    form = await request.form()
    try:
        plan_in = LocationPlanCreate(
            location=str(form.get("location") or ""),
            participants=int(form.get("participants") or 1),
            stock_duration_days=int(form.get("stock_duration_days") or 1),
        )
    except (ValueError, Exception) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    plan.location = plan_in.location
    plan.participants = plan_in.participants
    plan.stock_duration_days = plan_in.stock_duration_days
    plan.updated_at = datetime.now(UTC)
    session.add(plan)
    session.commit()
    return RedirectResponse("/location-plans?m=location-plan-saved", status_code=303)


@app.post("/location-plans/{plan_id}/delete")
async def location_plans_delete(
    plan_id: int,
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    plan = session.get(LocationPlan, plan_id)
    if plan:
        session.delete(plan)
        session.commit()
    return RedirectResponse("/location-plans?m=location-plan-deleted", status_code=303)


@app.get("/device-check")
def device_check(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    return _render(request, "device_check.html", {"user": maybe_user})


@app.get("/food-wheel")
def food_wheel_page(
    request: Request,
    location: str | None = None,
    session: Session = Depends(get_session),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user

    import logging as _logging
    _fw_log = _logging.getLogger("stockmgr.food_wheel")

    try:
        # All locations for filter dropdown (unfiltered)
        all_location_rows = session.exec(
            select(StockItem.storage_location).distinct()
        ).all()
        all_locations: list[str] = sorted(loc for loc in all_location_rows if loc)

        query = select(
            StockItem.food_group,
            StockItem.storage_location,
            StockItem.quantity,
            StockItem.unidose_per_pack,
        ).where(
            or_(StockItem.item_category == "food", StockItem.item_category == None)  # noqa: E711
        )
        if location:
            query = query.where(StockItem.storage_location == location)
        all_rows = session.exec(query).all()

        items_for_chart = [
            {"food_group": row[0], "quantity": row[2], "unidose_per_pack": row[3]}
            for row in all_rows
        ]
        lang = request.session.get("language", "pt")
        ungrouped_count = sum(1 for item in items_for_chart if not item["food_group"])

        try:
            chart = food_group_chart_data(items_for_chart, language=lang)
            chart_data = [
                {
                    "label": g["label"],
                    "color": g["color"],
                    "unidoses": int(g["actual_unidoses"]),
                    "actual_pct": g["actual_pct"],
                    "target_pct": g["target_pct"],
                    "delta": g["delta_pct"],
                }
                for g in chart["group_stats"]
            ]
            chart_total = chart["total_unidoses"]
            chart_json = json.dumps(chart_data)
        except Exception:
            _fw_log.exception("food_wheel chart build error")
            chart_data = []
            chart_total = 0
            chart_json = "[]"

        # Build plan tabs: convert to plain dicts to avoid DetachedInstanceError
        # (session closes before Jinja2 renders; ORM objects expire on session close)
        plans = session.exec(select(LocationPlan).order_by(LocationPlan.location)).all()
        plan_tabs = []
        for plan in plans:
            plan_items = session.exec(
                select(StockItem)
                .where(
                    StockItem.storage_location == plan.location,
                    StockItem.target_unidoses_location > 0,
                )
                .order_by(StockItem.name)
            ).all()
            plan_tabs.append({
                "plan": {
                    "location": plan.location,
                    "participants": plan.participants,
                    "stock_duration_days": plan.stock_duration_days,
                    "total_meal_occasions": plan.total_meal_occasions,
                },
                "plan_items": [
                    {
                        "name": item.name,
                        "item_type": item.item_type,
                        "target_unidoses_location": item.target_unidoses_location,
                        "unidose_per_pack": item.unidose_per_pack,
                    }
                    for item in plan_items
                ],
            })

    except Exception:
        _fw_log.exception("food_wheel_page unexpected error")
        raise HTTPException(status_code=500, detail="Food wheel error — check server logs.")

    return _render(
        request,
        "food_wheel.html",
        {
            "user": user,
            "chart_data": chart_data if chart_total > 0 else [],
            "chart_json": chart_json,
            "ungrouped_count": ungrouped_count,
            "all_locations": all_locations,
            "selected_location": location or "",
            "plan_tabs": plan_tabs,
        },
    )


@app.get("/renewals")
def renewal_plan(
    request: Request,
    session: Session = Depends(get_session),
    days: int | None = None,
    location: str | None = None,
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user

    window_days = days if days and days > 0 else settings.renewal_window_days
    start = date.today()
    end = start + timedelta(days=window_days)

    all_location_rows = session.exec(
        select(StockItem.storage_location).distinct()
    ).all()
    all_locations: list[str] = sorted(loc for loc in all_location_rows if loc)

    statement = (
        select(StockItem)
        .where(
            StockItem.expiry_date.is_not(None),
            StockItem.expiry_date >= start,
            StockItem.expiry_date <= end,
        )
        .order_by(StockItem.expiry_date, StockItem.name)
    )
    if location:
        statement = statement.where(StockItem.storage_location == location)
    rows = session.exec(statement).all()
    return _render(
        request,
        "renewals.html",
        {
            "user": user,
            "window_days": window_days,
            "rows": rows,
            "all_locations": all_locations,
            "selected_location": location or "",
        },
    )


@app.get("/products/by-name/{item_type}/{product_name:path}")
def product_detail(
    item_type: str,
    product_name: str,
    request: Request,
    session: Session = Depends(get_session),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user

    batches = _product_batches(
        session,
        item_type=item_type,
        product_name=product_name,
    )
    if not batches:
        raise HTTPException(status_code=404, detail="Product not found.")

    primary_batch = batches[0]
    product_summary = {
        "barcode": primary_batch.barcode,
        "temp_min_c": primary_batch.temp_min_c,
        "temp_max_c": primary_batch.temp_max_c,
        "humidity_min_pct": primary_batch.humidity_min_pct,
        "humidity_max_pct": primary_batch.humidity_max_pct,
        "renewal_date": primary_batch.renewal_date,
        "image_url": primary_batch.image_url,
        "nutriscore": primary_batch.nutriscore,
        "batch_count": len(batches),
        "total_quantity": sum(batch.quantity for batch in batches),
        "total_unidoses": sum(batch.quantity * batch.unidose_per_pack for batch in batches),
    }
    location_options = sorted(
        {batch.storage_location for batch in batches if batch.storage_location}
    )

    movement_query = (
        select(StockMovement, StockItem)
        .join(StockItem, StockItem.id == StockMovement.stock_item_id)
        .where(
            StockItem.item_type == item_type,
            StockItem.name == product_name,
        )
        .order_by(StockMovement.created_at.desc())
    )
    movement_rows = session.exec(movement_query).all()

    # Build per-location plan info map
    location_plans: dict[str, dict] = {}
    for loc in location_options:
        plan = session.exec(select(LocationPlan).where(LocationPlan.location == loc)).first()
        if plan:
            location_plans[loc] = {
                "participants": plan.participants,
                "days": plan.stock_duration_days,
                "total_meal_occasions": plan.total_meal_occasions,
            }

    return _render(
        request,
        "product_detail.html",
        {
            "user": user,
            "item_type": item_type,
            "product_name": product_name,
            "edit_item_id": primary_batch.id,
            "product_summary": product_summary,
            "location_options": location_options,
            "batches": batches,
            "movement_rows": movement_rows,
            "location_plans": location_plans,
            "message": _fetch_message(request),
        },
    )


@app.get("/items/new")
def item_new(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    prefill_name = request.query_params.get("name", "").strip()
    prefill_barcode = request.query_params.get("barcode", "").replace(" ", "").strip()
    prefill_item_type = request.query_params.get("item_type", "").strip()
    prefill_location = request.query_params.get("storage_location", "").strip()
    prefill_bucket = request.query_params.get("storage_bucket", "").strip()
    plans_map = {p.location: p.total_meal_occasions for p in session.exec(select(LocationPlan)).all()}
    return _render(
        request,
        "item_form.html",
        {
            "user": maybe_user,
            "mode": "create",
            "draft": {
                "name": prefill_name,
                "item_type": prefill_item_type or "unknown",
                "barcode": prefill_barcode,
                "storage_location": prefill_location,
                "storage_bucket": prefill_bucket,
                "quantity": 0,
                "unidose_per_pack": 1,
                "target_unidoses_location": 0,
            },
            "lookup": None,
            "food_groups": FOOD_GROUPS,
            "plan_locations": _plan_locations(session),
            "location_plans_json": json.dumps(plans_map),
            "lookup_providers": _lookup_provider_names(),
        },
    )


@app.get("/items/{item_id}/edit")
def item_edit(item_id: int, request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    item = session.get(StockItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found.")
    all_batches = _product_batches(
        session,
        item_type=item.item_type,
        product_name=item.name,
    )
    plans_map = {p.location: p.total_meal_occasions for p in session.exec(select(LocationPlan)).all()}
    back_url = f"/products/by-name/{item.item_type}/{item.name.replace(' ', '%20')}"
    return _render(
        request,
        "item_form.html",
        {
            "user": user,
            "mode": "edit",
            "draft": item,
            "lookup": None,
            "all_batches": all_batches,
            "food_groups": FOOD_GROUPS,
            "plan_locations": _plan_locations(session),
            "location_plans_json": json.dumps(plans_map),
            "back_url": back_url,
        },
    )


@app.get("/items/lookup-json")
async def lookup_json(
    request: Request,
    barcode: str,
    item_type: str = "unknown",
    session: Session = Depends(get_session),
):
    """AJAX barcode lookup — returns enriched JSON for the new-item form (no page refresh)."""
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        from fastapi.responses import JSONResponse
        return JSONResponse({"error": "unauthenticated"}, status_code=401)

    result = await barcode_service.lookup(barcode=barcode, item_type=item_type)
    payload: dict[str, Any] = {
        "found": result.found,
        "provider": result.provider,
        "attempts": [
            {"provider": a.provider, "status": a.status}
            for a in (result.attempts or [])
        ],
    }
    if result.found and result.data:
        inferred_group = infer_food_group(
            name=result.data.get("name") or "",
            item_type=item_type if item_type != "unknown" else result.data.get("category") or "",
            food_groups_tags=result.data.get("foodGroupsTags"),
        )
        wc, uom = _parse_size(result.data.get("size"))
        payload.update({
            "barcode": barcode,
            "name": result.data.get("name"),
            "item_type": item_type if item_type != "unknown" else (result.data.get("category") or "unknown"),
            "image_url": result.data.get("imageUrl"),
            "nutriscore": result.data.get("nutriscore"),
            "food_group": inferred_group,
            "weight_capacity": wc,
            "uom": uom,
        })
    from fastapi.responses import JSONResponse
    return JSONResponse(payload)


@app.post("/items/lookup")
async def lookup_for_form(
    request: Request,
    barcode: str = Form(...),
    item_type: str = Form("unknown"),
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    result = await barcode_service.lookup(barcode=barcode, item_type=item_type)
    draft: dict[str, Any] = {"barcode": barcode, "item_type": item_type, "quantity": 0}
    if result.found and result.data:
        inferred_group = infer_food_group(
            name=result.data.get("name") or "",
            item_type=item_type if item_type != "unknown" else result.data.get("category") or "",
            food_groups_tags=result.data.get("foodGroupsTags"),
        )
        draft.update(
            {
                "name": result.data.get("name"),
                "item_type": item_type
                if item_type != "unknown"
                else (result.data.get("category") or "unknown"),
                "batch_code": "",
                "storage_location": "",
                "storage_bucket": "",
                "unidose_per_pack": 1,
                "target_unidoses_location": 0,
                "image_url": result.data.get("imageUrl"),
                "nutriscore": result.data.get("nutriscore"),
                "food_group": inferred_group,
                **dict(zip(("weight_capacity", "uom"), _parse_size(result.data.get("size")))),
            }
        )
    location_context = _storage_location_field_context(
        session,
        selected_location=str(draft.get("storage_location") or ""),
    )
    plans_map = {p.location: p.total_meal_occasions for p in session.exec(select(LocationPlan)).all()}
    return _render(
        request,
        "item_form.html",
        {
            "user": maybe_user,
            "mode": "create",
            "draft": draft,
            "lookup": result,
            "food_groups": FOOD_GROUPS,
            "plan_locations": _plan_locations(session),
            "location_plans_json": json.dumps(plans_map),
            "lookup_providers": _lookup_provider_names(),
            **location_context,
        },
    )


@app.post("/items")
async def item_create(
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    form = await request.form()

    # Multi-location create: loc_location[] submitted per row
    loc_locations = form.getlist("loc_location")
    if loc_locations:
        # Filter to non-empty location rows
        loc_batch_codes = form.getlist("loc_batch_code")
        loc_expiries = form.getlist("loc_expiry")
        loc_quantities = form.getlist("loc_quantity")
        loc_buckets = form.getlist("loc_bucket")
        loc_renewals = form.getlist("loc_renewal")
        loc_targets = form.getlist("loc_target")

        base_payload = _item_payload_from_form(dict(form))
        if not base_payload.get("food_group"):
            base_payload["food_group"] = infer_food_group(
                name=base_payload.get("name") or "",
                item_type=base_payload.get("item_type") or "",
            )

        created_items: list[StockItem] = []
        for idx, loc in enumerate(loc_locations):
            loc = loc.strip()
            if not loc:
                continue
            expiry = loc_expiries[idx] if idx < len(loc_expiries) else ""
            if not expiry:
                continue
            row = dict(base_payload)
            row["storage_location"] = loc
            row["batch_code"] = (loc_batch_codes[idx] if idx < len(loc_batch_codes) else None) or None
            row["expiry_date"] = expiry
            row["quantity"] = int(loc_quantities[idx] or 0) if idx < len(loc_quantities) else 0
            row["storage_bucket"] = loc_buckets[idx] if idx < len(loc_buckets) else ""
            renewal_raw = loc_renewals[idx] if idx < len(loc_renewals) else ""
            row["renewal_date"] = renewal_raw if renewal_raw else None
            target_raw = loc_targets[idx] if idx < len(loc_targets) else 0
            row["target_unidoses_location"] = int(target_raw or 0)
            if not row["target_unidoses_location"]:
                plan = session.exec(select(LocationPlan).where(LocationPlan.location == loc)).first()
                if plan and row.get("food_group"):
                    fg = FOOD_GROUP_BY_KEY.get(row["food_group"])
                    if fg:
                        row["target_unidoses_location"] = math.ceil(
                            plan.total_meal_occasions * fg.target_pct / 100
                        )
            try:
                item_in = ItemCreate.model_validate(row)
            except ValidationError as exc:
                raise HTTPException(status_code=422, detail=exc.errors()) from exc
            item = StockItem(**item_in.model_dump(), user_id=user.id)
            session.add(item)
            session.commit()
            session.refresh(item)
            created_items.append(item)
            if item.quantity > 0:
                session.add(StockMovement(
                    stock_item_id=item.id,
                    user_id=user.id,
                    delta=item.quantity,
                    note="Initial stock quantity",
                ))
                session.commit()
            try:
                await calendar_service.schedule_renewal(user=user, item=item)
            except CalendarSyncError as exc:
                raise HTTPException(status_code=502, detail=str(exc)) from exc
        names = ", ".join(i.storage_location for i in created_items)
        await _notify_telegram_operation_async(
            operation="item-created",
            actor=user.email,
            detail=(
                f"name={base_payload.get('name')} type={base_payload.get('item_type')} "
                f"locations={names}"
            ),
        )
        # Auto-create location plans for any new locations submitted (#53)
        new_plan_locs = form.getlist("new_plan_location")
        new_plan_parts = form.getlist("new_plan_participants")
        new_plan_days = form.getlist("new_plan_days")
        for i, plan_loc in enumerate(new_plan_locs):
            plan_loc = plan_loc.strip()
            if not plan_loc:
                continue
            try:
                parts = int(new_plan_parts[i]) if i < len(new_plan_parts) else 4
                days = int(new_plan_days[i]) if i < len(new_plan_days) else 30
                _upsert_location_plan(session, location=plan_loc, participants=parts, stock_duration_days=days)
            except (ValueError, IndexError):
                pass
        if form.get("continue_adding"):
            return RedirectResponse("/items/new?m=item-created", status_code=303)
        return RedirectResponse("/?m=item-created", status_code=303)

    # Legacy single-row path
    payload = _item_payload_from_form(dict(form))
    # Auto-infer food_group from name/type if not set
    if not payload.get("food_group"):
        payload["food_group"] = infer_food_group(
            name=payload.get("name") or "",
            item_type=payload.get("item_type") or "",
        )
    # Auto-compute target_unidoses_location from location plan if not set
    if not payload.get("target_unidoses_location"):
        loc = payload.get("storage_location") or ""
        plan = session.exec(select(LocationPlan).where(LocationPlan.location == loc)).first()
        if plan and payload.get("food_group"):
            fg = FOOD_GROUP_BY_KEY.get(payload["food_group"])
            if fg:
                payload["target_unidoses_location"] = math.ceil(
                    plan.total_meal_occasions * fg.target_pct / 100
                )
    try:
        item_in = ItemCreate.model_validate(payload)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors()) from exc
    item = StockItem(**item_in.model_dump(), user_id=user.id)
    session.add(item)
    session.commit()
    session.refresh(item)

    if item.quantity > 0:
        session.add(
            StockMovement(
                stock_item_id=item.id,
                user_id=user.id,
                delta=item.quantity,
                note="Initial stock quantity",
            )
        )
        session.commit()

    try:
        await calendar_service.schedule_renewal(user=user, item=item)
    except CalendarSyncError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    await _notify_telegram_operation_async(
        operation="item-created",
        actor=user.email,
        detail=(
            f"name={item.name} type={item.item_type} qty={item.quantity} "
            f"location={item.storage_location}"
        ),
    )

    if form.get("new_plan_location"):
        try:
            _upsert_location_plan(
                session,
                location=str(form.get("new_plan_location")),
                participants=int(form.get("new_plan_participants") or 4),
                stock_duration_days=int(form.get("new_plan_days") or 30),
            )
        except (ValueError, Exception):
            pass

    if form.get("continue_adding"):
        return RedirectResponse("/items/new?m=item-created", status_code=303)
    return RedirectResponse("/?m=item-created", status_code=303)


@app.post("/items/{item_id}/update")
async def item_update(
    item_id: int,
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    item = session.get(StockItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found.")

    form = await request.form()

    # ── Multi-row edit path ──────────────────────────────────────────────
    row_ids = form.getlist("row_id")
    if row_ids is not None and len(row_ids) > 0:
        # Product-level fields (shared across all batches)
        product_payload = _item_payload_from_form(dict(form))
        if not product_payload.get("food_group"):
            product_payload["food_group"] = infer_food_group(
                name=product_payload.get("name") or "",
                item_type=product_payload.get("item_type") or "",
            )

        # Delete marked batches (safety: only delete batches for this product)
        delete_ids_raw = form.getlist("delete_ids")
        for raw_id in delete_ids_raw:
            try:
                del_id = int(raw_id)
            except (ValueError, TypeError):
                continue
            del_item = session.get(StockItem, del_id)
            if del_item and del_item.name == item.name and del_item.item_type == item.item_type:
                session.delete(del_item)
        session.commit()

        row_locations = form.getlist("row_location")
        row_batch_codes = form.getlist("row_batch_code")
        row_expiries = form.getlist("row_expiry")
        row_quantities = form.getlist("row_quantity")
        row_buckets = form.getlist("row_bucket")
        row_targets = form.getlist("row_target")
        row_renewals = form.getlist("row_renewal")

        for idx, raw_row_id in enumerate(row_ids):
            loc = (row_locations[idx] if idx < len(row_locations) else "").strip()
            if not loc:
                continue
            expiry = row_expiries[idx] if idx < len(row_expiries) else ""
            if not expiry:
                continue

            row = dict(product_payload)
            row["storage_location"] = loc
            row["batch_code"] = (row_batch_codes[idx] if idx < len(row_batch_codes) else None) or None
            row["expiry_date"] = expiry
            row["quantity"] = int(row_quantities[idx] or 0) if idx < len(row_quantities) else 0
            row["storage_bucket"] = row_buckets[idx] if idx < len(row_buckets) else ""
            renewal_raw = row_renewals[idx] if idx < len(row_renewals) else ""
            row["renewal_date"] = renewal_raw if renewal_raw else None
            target_raw = row_targets[idx] if idx < len(row_targets) else 0
            row["target_unidoses_location"] = int(target_raw or 0)
            if not row["target_unidoses_location"]:
                plan = session.exec(select(LocationPlan).where(LocationPlan.location == loc)).first()
                if plan and row.get("food_group"):
                    fg = FOOD_GROUP_BY_KEY.get(row["food_group"])
                    if fg:
                        row["target_unidoses_location"] = math.ceil(
                            plan.total_meal_occasions * fg.target_pct / 100
                        )

            try:
                item_in = ItemCreate.model_validate(row)
            except ValidationError as exc:
                raise HTTPException(status_code=422, detail=exc.errors()) from exc

            raw_row_id = raw_row_id.strip()
            if raw_row_id:
                existing = session.get(StockItem, int(raw_row_id))
                if existing and existing.name == item.name and existing.item_type == item.item_type:
                    previous_qty = existing.quantity
                    for key, value in item_in.model_dump().items():
                        setattr(existing, key, value)
                    existing.updated_at = datetime.now(UTC)
                    session.add(existing)
                    session.commit()
                    session.refresh(existing)
                    qty_delta = existing.quantity - previous_qty
                    if qty_delta:
                        session.add(StockMovement(
                            stock_item_id=existing.id,
                            user_id=user.id,
                            delta=qty_delta,
                            note="Quantity updated in edit form",
                        ))
                    try:
                        await calendar_service.schedule_renewal(user=user, item=existing)
                    except CalendarSyncError as exc:
                        raise HTTPException(status_code=502, detail=str(exc)) from exc
            else:
                new_item = StockItem(**item_in.model_dump(), user_id=user.id)
                session.add(new_item)
                session.commit()
                session.refresh(new_item)
                if new_item.quantity > 0:
                    session.add(StockMovement(
                        stock_item_id=new_item.id,
                        user_id=user.id,
                        delta=new_item.quantity,
                        note="Initial stock quantity",
                    ))
                try:
                    await calendar_service.schedule_renewal(user=user, item=new_item)
                except CalendarSyncError as exc:
                    raise HTTPException(status_code=502, detail=str(exc)) from exc

        session.commit()
        await _notify_telegram_operation_async(
            operation="item-updated",
            actor=user.email,
            detail=f"name={item.name} type={item.item_type} rows={len(row_ids)}",
        )
        product_name_enc = item.name.replace(" ", "%20")
        return RedirectResponse(
            f"/products/by-name/{item.item_type}/{product_name_enc}?m=item-updated",
            status_code=303,
        )

    # ── Legacy single-field path (API / backward-compat) ────────────────
    previous_quantity = item.quantity
    payload = _item_payload_from_form(dict(form))
    if not payload.get("food_group"):
        payload["food_group"] = infer_food_group(
            name=payload.get("name") or "",
            item_type=payload.get("item_type") or "",
        )
    if not payload.get("target_unidoses_location"):
        loc = payload.get("storage_location") or ""
        plan = session.exec(select(LocationPlan).where(LocationPlan.location == loc)).first()
        if plan and payload.get("food_group"):
            fg = FOOD_GROUP_BY_KEY.get(payload["food_group"])
            if fg:
                payload["target_unidoses_location"] = math.ceil(
                    plan.total_meal_occasions * fg.target_pct / 100
                )
    try:
        item_in = ItemCreate.model_validate(payload)
    except ValidationError as exc:
        raise HTTPException(status_code=422, detail=exc.errors()) from exc

    for key, value in item_in.model_dump().items():
        setattr(item, key, value)
    item.updated_at = datetime.now(UTC)
    session.add(item)
    session.commit()
    session.refresh(item)

    quantity_delta = item.quantity - previous_quantity
    if quantity_delta:
        session.add(
            StockMovement(
                stock_item_id=item.id,
                user_id=user.id,
                delta=quantity_delta,
                note="Quantity updated in edit form",
            )
        )
        session.commit()

    try:
        await calendar_service.schedule_renewal(user=user, item=item)
    except CalendarSyncError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    await _notify_telegram_operation_async(
        operation="item-updated",
        actor=user.email,
        detail=(
            f"name={item.name} type={item.item_type} qty={item.quantity} "
            f"delta={quantity_delta} location={item.storage_location}"
        ),
    )

    return RedirectResponse("/?m=item-updated", status_code=303)


@app.post("/items/{item_id}/move")
def adjust_stock_quantity(
    item_id: int,
    request: Request,
    direction: str = Form(...),
    quantity_step: int = Form(1),
    note: str = Form(""),
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    item = session.get(StockItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found.")

    step = max(1, int(quantity_step))
    if direction not in {"in", "out"}:
        raise HTTPException(status_code=400, detail="Invalid stock movement direction.")
    delta = step if direction == "in" else -step
    new_quantity = item.quantity + delta
    if new_quantity < 0:
        raise HTTPException(status_code=400, detail="Cannot reduce below zero.")

    item.quantity = new_quantity
    item.updated_at = datetime.now(UTC)
    session.add(item)
    session.add(
        StockMovement(
            stock_item_id=item.id,
            user_id=user.id,
            delta=delta,
            note=(note or "").strip() or None,
        )
    )
    session.commit()
    _notify_telegram_operation_sync(
        operation="stock-move",
        actor=user.email,
        detail=(
            f"name={item.name} type={item.item_type} delta={delta} "
            f"new_qty={item.quantity} note={(note or '').strip() or '-'}"
        ),
    )

    target = request.headers.get("referer") or "/"
    return RedirectResponse(target, status_code=303)


@app.post("/items/{item_id}/delete")
def item_delete(
    item_id: int,
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    item = session.get(StockItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found.")
    deleted_name = item.name
    deleted_type = item.item_type
    deleted_location = item.storage_location
    session.delete(item)
    session.commit()
    _notify_telegram_operation_sync(
        operation="item-deleted",
        actor=user.email,
        detail=f"name={deleted_name} type={deleted_type} location={deleted_location}",
    )
    return RedirectResponse("/?m=item-deleted", status_code=303)




@app.get("/items/export")
def export_items(request: Request, session: Session = Depends(get_session)):
    """Export all stock items to XLSX in import-compatible format."""
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user

    from io import BytesIO
    from openpyxl import Workbook

    items = session.exec(select(StockItem).order_by(StockItem.name, StockItem.storage_location)).all()

    columns = [
        ("barcode", "barcode"),
        ("batch_code", "batch_code"),
        ("name", "name"),
        ("item_type", "item_type"),
        ("storage_location", "storage_location"),
        ("storage_bucket", "storage_bucket"),
        ("expiry_date", "expiry_date"),
        ("quantity", "quantity"),
        ("unidose_per_pack", "unidose_per_pack"),
        ("target_unidoses_location", "target_unidoses_location"),
        ("weight_capacity", "weight_capacity"),
        ("uom", "uom"),
        ("temp_min_c", "temp_min_c"),
        ("temp_max_c", "temp_max_c"),
        ("humidity_min_pct", "humidity_min_pct"),
        ("humidity_max_pct", "humidity_max_pct"),
        ("renewal_date", "renewal_date"),
        ("comment", "comment"),
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append([col[0] for col in columns])
    for item in items:
        ws.append([getattr(item, col[1], None) for col in columns])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import Response
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=stockmgr-export.xlsx"},
    )


@app.post("/items/{item_id}/upload-image")
async def upload_item_image(
    item_id: int,
    request: Request,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    """Upload a product image for a stock item."""
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user

    item = session.get(StockItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Item not found.")

    _allowed_mime = frozenset({"image/jpeg", "image/png", "image/webp", "image/gif"})
    if file.content_type not in _allowed_mime:
        raise HTTPException(status_code=422, detail="Only JPEG, PNG, WebP and GIF images are allowed.")

    # Filename is fully server-controlled (pure random token, no user-provided data)
    filename = f"img_{secrets.token_hex(24)}.jpg"
    dest = _MEDIA_DIR / filename
    dest.write_bytes(await file.read())

    item.image_url = f"/media/{filename}"
    item.updated_at = datetime.now(UTC)
    session.add(item)
    session.commit()
    safe_id = int(item_id)
    return RedirectResponse(f"/items/{safe_id}/edit?m=image-updated", status_code=303)


@app.get("/items/import")
def import_page(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    return _render(request, "import.html", {"user": maybe_user, "result": None})


@app.post("/items/import")
async def import_items(
    request: Request,
    file: UploadFile = File(...),
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user
    file_bytes = await file.read()
    items, result = parse_import_file(file_bytes, file.filename or "")
    for item_in in items:
        item = StockItem(**item_in.model_dump(), user_id=user.id)
        session.add(item)
    session.commit()
    await _notify_telegram_operation_async(
        operation="items-imported",
        actor=user.email,
        detail=f"imported={result.imported} failed={result.failed}",
    )
    return _render(request, "import.html", {"user": user, "result": result})


@app.get("/admin/users")
def admin_users(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
):
    users = session.exec(select(User).order_by(User.requested_at.desc())).all()
    return _render(
        request,
        "admin_users.html",
        {
            "user": admin,
            "users": users,
            "message": _fetch_message(request),
            "enriched_count": request.query_params.get("enriched"),
        },
    )


@app.post("/admin/users/{user_id}/approve")
def admin_approve_user(
    user_id: int,
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
    _csrf: None = Depends(_validate_csrf),
):
    _ = request
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    user.approval_status = "approved"
    user.approved_at = datetime.now(UTC)
    session.add(user)
    session.commit()
    return RedirectResponse("/admin/users?m=user-approved", status_code=303)


@app.post("/admin/users/{user_id}/reject")
def admin_reject_user(
    user_id: int,
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
    _csrf: None = Depends(_validate_csrf),
):
    _ = request
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    user.approval_status = "rejected"
    user.approved_at = None
    session.add(user)
    session.commit()
    return RedirectResponse("/admin/users?m=user-rejected", status_code=303)


@app.post("/admin/users/{user_id}/toggle-admin")
def admin_toggle_admin(
    user_id: int,
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
    _csrf: None = Depends(_validate_csrf),
):
    _ = request
    user = session.get(User, user_id)
    if not user:
        raise HTTPException(status_code=404, detail="User not found.")
    if user.id == admin.id and user.is_admin and _count_approved_admins(session) <= 1:
        raise HTTPException(status_code=400, detail="Cannot remove the last admin.")
    user.is_admin = not user.is_admin
    session.add(user)
    session.commit()
    return RedirectResponse("/admin/users?m=user-role-updated", status_code=303)




@app.get("/admin/enrich")
def admin_enrich_page(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
):
    _ = session
    return _render(
        request,
        "enrich.html",
        {
            "user": admin,
            "message": _fetch_message(request),
            "enriched_count": request.query_params.get("enriched"),
        },
    )


@app.get("/admin/enrich-stream")
async def admin_enrich_stream(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
):
    """SSE endpoint that streams per-item enrichment progress."""
    import asyncio

    async def _generate():
        items_with_barcode = session.exec(
            select(StockItem).where(
                StockItem.barcode.is_not(None),  # type: ignore[union-attr]
                StockItem.barcode != "",
            )
        ).all()
        total = len(items_with_barcode)
        enriched = 0
        for i, item in enumerate(items_with_barcode):
            skipped = item.image_url and item.nutriscore and item.food_group
            result_tag = "skipped"
            provider_name = ""
            if not skipped:
                try:
                    result = await barcode_service.lookup(
                        barcode=item.barcode, item_type=item.item_type
                    )
                    if result.found and result.data:
                        changed = False
                        if not item.image_url and result.data.get("imageUrl"):
                            item.image_url = result.data["imageUrl"]
                            changed = True
                        if not item.nutriscore and result.data.get("nutriscore"):
                            item.nutriscore = result.data["nutriscore"]
                            changed = True
                        if not item.food_group:
                            inferred = infer_food_group(
                                name=result.data.get("name") or item.name,
                                item_type=item.item_type,
                                food_groups_tags=result.data.get("foodGroupsTags"),
                            )
                            if inferred:
                                item.food_group = inferred
                                changed = True
                        if not item.weight_capacity and result.data.get("size"):
                            wc, u = _parse_size(result.data.get("size"))
                            if wc is not None:
                                item.weight_capacity = wc
                                item.uom = u
                                changed = True
                        if changed:
                            item.updated_at = datetime.now(UTC)
                            session.add(item)
                            session.commit()
                            enriched += 1
                            result_tag = "updated"
                            provider_name = result.provider or ""
                        else:
                            result_tag = "skipped"
                    else:
                        result_tag = "not_found"
                except Exception:
                    result_tag = "error"
            progress_data = json.dumps({
                "current": i + 1,
                "total": total,
                "name": item.name,
                "result": result_tag,
                "provider": provider_name,
            })
            yield f"data: {progress_data}\n\n"
            await asyncio.sleep(0)
        done_data = json.dumps({"done": True, "enriched": enriched})
        yield f"data: {done_data}\n\n"

    return StreamingResponse(
        _generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/admin/enrich-items")
async def admin_enrich_items(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
    _csrf: None = Depends(_validate_csrf),
):
    """Re-query barcode providers for all items missing image_url or nutriscore."""
    _ = request
    items_with_barcode = session.exec(
        select(StockItem).where(
            StockItem.barcode.is_not(None),  # type: ignore[union-attr]
            StockItem.barcode != "",
        )
    ).all()
    enriched = 0
    for item in items_with_barcode:
        if item.image_url and item.nutriscore and item.food_group:
            continue
        try:
            result = await barcode_service.lookup(
                barcode=item.barcode, item_type=item.item_type
            )
        except Exception:
            continue
        if not result.found or not result.data:
            continue
        changed = False
        if not item.image_url and result.data.get("imageUrl"):
            item.image_url = result.data["imageUrl"]
            changed = True
        if not item.nutriscore and result.data.get("nutriscore"):
            item.nutriscore = result.data["nutriscore"]
            changed = True
        if not item.food_group:
            inferred = infer_food_group(
                name=result.data.get("name") or item.name,
                item_type=item.item_type,
                food_groups_tags=result.data.get("foodGroupsTags"),
            )
            if inferred:
                item.food_group = inferred
                changed = True
        if not item.weight_capacity and result.data.get("size"):
            wc, u = _parse_size(result.data.get("size"))
            if wc is not None:
                item.weight_capacity = wc
                item.uom = u
                changed = True
        if changed:
            item.updated_at = datetime.now(UTC)
            session.add(item)
            enriched += 1
    session.commit()
    return RedirectResponse(f"/admin/enrich?m=enrich-done&enriched={enriched}", status_code=303)


@app.get("/admin/info")
def admin_info_page(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
):
    import os
    from sqlmodel import text as sql_text  # noqa: F401

    db_size_str = "N/A"
    db_url = settings.database_url
    if db_url.startswith("sqlite"):
        try:
            from app.db import _sqlite_path_from_url
            db_path = _sqlite_path_from_url(db_url)
            if db_path:
                if not db_path.is_absolute():
                    db_path = (Path.cwd() / db_path).resolve()
                size_bytes = os.path.getsize(db_path)
                if size_bytes >= 1_048_576:
                    db_size_str = f"{size_bytes / 1_048_576:.1f} MB"
                else:
                    db_size_str = f"{size_bytes / 1024:.1f} KB"
        except Exception:
            pass

    db_item_count = session.exec(select(func.count()).select_from(StockItem)).one()
    db_user_count = session.exec(select(func.count()).select_from(User)).one()

    return _render(
        request,
        "admin_info.html",
        {
            "user": admin,
            "runtime_sha": settings.app_version,
            "db_size_str": db_size_str,
            "db_item_count": db_item_count,
            "db_user_count": db_user_count,
        },
    )


@app.get("/admin/backup")
def admin_backup(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
):
    import openpyxl

    _ = request
    items = session.exec(select(StockItem)).all()
    plans = session.exec(select(LocationPlan)).all()

    # Build stock_items.xlsx
    wb_items = openpyxl.Workbook()
    ws_items = wb_items.active
    ws_items.append([
        "barcode", "batch_code", "name", "item_type", "storage_location",
        "storage_bucket", "expiry_date", "quantity", "unidose_per_pack",
        "target_unidoses_location", "weight_capacity", "uom",
        "temp_min_c", "temp_max_c", "humidity_min_pct", "humidity_max_pct",
        "renewal_date", "comment", "image_url",
    ])
    for item in items:
        ws_items.append([
            item.barcode, item.batch_code, item.name, item.item_type,
            item.storage_location, item.storage_bucket,
            item.expiry_date.isoformat() if item.expiry_date else None,
            item.quantity, item.unidose_per_pack, item.target_unidoses_location,
            item.weight_capacity, item.uom, item.temp_min_c, item.temp_max_c,
            item.humidity_min_pct, item.humidity_max_pct,
            item.renewal_date.isoformat() if item.renewal_date else None,
            item.comment, item.image_url,
        ])
    buf_items = io.BytesIO()
    wb_items.save(buf_items)
    items_xlsx_bytes = buf_items.getvalue()

    # Build location_plans.xlsx
    wb_plans = openpyxl.Workbook()
    ws_plans = wb_plans.active
    ws_plans.append(["location", "participants", "stock_duration_days"])
    for plan in plans:
        ws_plans.append([plan.location, plan.participants, plan.stock_duration_days])
    buf_plans = io.BytesIO()
    wb_plans.save(buf_plans)
    plans_xlsx_bytes = buf_plans.getvalue()

    # Build metadata.json
    metadata = {
        "version": settings.app_version,
        "backup_date": datetime.now(UTC).isoformat(),
        "item_count": len(items),
        "plan_count": len(plans),
    }

    # Pack into ZIP
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("metadata.json", json.dumps(metadata))
        zf.writestr("stock_items.xlsx", items_xlsx_bytes)
        zf.writestr("location_plans.xlsx", plans_xlsx_bytes)
    zip_buffer.seek(0)

    date_str = datetime.now(UTC).strftime("%Y-%m-%d")
    return StreamingResponse(
        zip_buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename=stockmgr-backup-{date_str}.zip"},
    )


@app.get("/admin/restore")
def admin_restore_page(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
):
    _ = session
    return _render(request, "restore.html", {"user": admin, "result": None})


@app.post("/admin/restore")
async def admin_restore(
    request: Request,
    session: Session = Depends(get_session),
    admin: User = Depends(_require_admin_user),
    _csrf: None = Depends(_validate_csrf),
    file: UploadFile = File(...),
):
    import openpyxl

    file_bytes = await file.read()
    try:
        zf = zipfile.ZipFile(io.BytesIO(file_bytes))
    except zipfile.BadZipFile:
        raise HTTPException(status_code=400, detail="Invalid backup file: not a ZIP archive")

    if "metadata.json" not in zf.namelist():
        raise HTTPException(
            status_code=400,
            detail=translate(_current_language(request), "msg.backup-restore-invalid"),
        )

    items_imported = 0
    if "stock_items.xlsx" in zf.namelist():
        xlsx_bytes = zf.read("stock_items.xlsx")
        parsed_items, _ = parse_import_file(xlsx_bytes, "stock_items.xlsx")
        for item_create in parsed_items:
            existing = session.exec(
                select(StockItem).where(
                    StockItem.name == item_create.name,
                    StockItem.storage_location == item_create.storage_location,
                    StockItem.batch_code == item_create.batch_code,
                )
            ).first()
            if existing:
                for field in item_create.model_fields:
                    val = getattr(item_create, field, None)
                    if val is not None:
                        setattr(existing, field, val)
                existing.updated_at = datetime.now(UTC)
                session.add(existing)
            else:
                new_item = StockItem(**item_create.model_dump(), user_id=admin.id)
                session.add(new_item)
            items_imported += 1

    plans_imported = 0
    if "location_plans.xlsx" in zf.namelist():
        xlsx_bytes = zf.read("location_plans.xlsx")
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            for row in rows[1:]:  # skip header
                if not row or row[0] is None:
                    continue
                loc, participants, duration = str(row[0]), row[1], row[2]
                if participants is None or duration is None:
                    continue
                existing_plan = session.exec(
                    select(LocationPlan).where(LocationPlan.location == loc)
                ).first()
                if existing_plan:
                    existing_plan.participants = int(participants)
                    existing_plan.stock_duration_days = int(duration)
                    existing_plan.updated_at = datetime.now(UTC)
                    session.add(existing_plan)
                else:
                    new_plan = LocationPlan(
                        location=loc,
                        participants=int(participants),
                        stock_duration_days=int(duration),
                    )
                    session.add(new_plan)
                plans_imported += 1

    session.commit()
    return _render(
        request,
        "restore.html",
        {
            "user": admin,
            "result": {"items_imported": items_imported, "plans_imported": plans_imported},
        },
    )


@app.get("/admin/config")
def admin_config_page(
    request: Request,
    admin: User = Depends(_require_admin_user),
):
    config_path = Path(settings.provider_config_path)
    if not config_path.is_absolute():
        config_path = Path(__file__).resolve().parents[1] / settings.provider_config_path
    config_json = config_path.read_text(encoding="utf-8")
    settings_dict = {
        "environment": settings.environment,
        "app_version": settings.app_version,
        "auth_mode": settings.auth_mode,
        "calendar_provider": settings.calendar_provider,
        "renewal_window_days": settings.renewal_window_days,
        "admin_emails": settings.admin_emails,
        "provider_config_path": settings.provider_config_path,
        "provider_schema_path": settings.provider_schema_path,
    }
    return _render(
        request,
        "config.html",
        {
            "user": admin,
            "config_json": json.dumps(json.loads(config_json), indent=2),
            "settings_dict": settings_dict,
            "message": _fetch_message(request),
            "error": None,
        },
    )


@app.post("/admin/config")
async def admin_config_save(
    request: Request,
    admin: User = Depends(_require_admin_user),
    _csrf: None = Depends(_validate_csrf),
    config_json: str = Form(...),
):
    settings_dict = {
        "environment": settings.environment,
        "app_version": settings.app_version,
        "auth_mode": settings.auth_mode,
        "calendar_provider": settings.calendar_provider,
        "renewal_window_days": settings.renewal_window_days,
        "admin_emails": settings.admin_emails,
        "provider_config_path": settings.provider_config_path,
        "provider_schema_path": settings.provider_schema_path,
    }
    try:
        parsed = json.loads(config_json)
    except json.JSONDecodeError as exc:
        return _render(
            request,
            "config.html",
            {
                "user": admin,
                "config_json": config_json,
                "settings_dict": settings_dict,
                "message": None,
                "error": f"{translate(_current_language(request), 'config.validation_error')}: {exc}",
            },
        )

    schema_path = Path(settings.provider_schema_path)
    if not schema_path.is_absolute():
        schema_path = Path(__file__).resolve().parents[1] / settings.provider_schema_path
    schema = json.loads(schema_path.read_text(encoding="utf-8"))

    try:
        jsonschema_validate(instance=parsed, schema=schema)
    except JsonSchemaValidationError as exc:
        return _render(
            request,
            "config.html",
            {
                "user": admin,
                "config_json": config_json,
                "settings_dict": settings_dict,
                "message": None,
                "error": f"{translate(_current_language(request), 'config.validation_error')}: {exc.message}",
            },
        )

    config_path = Path(settings.provider_config_path)
    if not config_path.is_absolute():
        config_path = Path(__file__).resolve().parents[1] / settings.provider_config_path
    config_path.write_text(json.dumps(parsed, indent=2), encoding="utf-8")
    barcode_service.reload_config()

    return RedirectResponse("/admin/config?m=saved", status_code=303)



@app.get("/items/unidose-plan")
def unidose_plan_page(request: Request, session: Session = Depends(get_session)):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user
    user = maybe_user

    all_plans = session.exec(select(LocationPlan)).all()
    plans_by_location: dict[str, LocationPlan] = {p.location: p for p in all_plans}

    all_items = session.exec(select(StockItem).order_by(StockItem.name)).all()
    seen: dict[str, dict] = {}
    for item in all_items:
        if item.name in seen:
            continue
        plan = plans_by_location.get(item.storage_location)
        if plan and plan.participants > 0 and plan.stock_duration_days > 0:
            denom = plan.participants * plan.stock_duration_days
            current_upd = round(item.target_unidoses_location / denom, 2)
            has_plan = True
        else:
            current_upd = 0.0
            has_plan = False
        seen[item.name] = {
            "name": item.name,
            "item_type": item.item_type,
            "barcode": item.barcode or "",
            "current_unidose_per_day": current_upd,
            "has_plan": has_plan,
        }
    products = list(seen.values())

    return _render(request, "unidose_plan.html", {
        "user": user,
        "products": products,
        "message": request.query_params.get("m"),
    })


@app.post("/items/unidose-plan")
async def unidose_plan_submit(
    request: Request,
    session: Session = Depends(get_session),
    _csrf: None = Depends(_validate_csrf),
):
    maybe_user = _require_user_or_redirect(request, session)
    if isinstance(maybe_user, RedirectResponse):
        return maybe_user

    all_plans = session.exec(select(LocationPlan)).all()
    plans_by_location: dict[str, LocationPlan] = {p.location: p for p in all_plans}

    form = await request.form()

    all_items = session.exec(select(StockItem)).all()
    old_upd: dict[str, float] = {}
    for item in all_items:
        if item.name in old_upd:
            continue
        plan = plans_by_location.get(item.storage_location)
        if plan and plan.participants > 0 and plan.stock_duration_days > 0:
            denom = plan.participants * plan.stock_duration_days
            old_upd[item.name] = round(item.target_unidoses_location / denom, 2)
        else:
            old_upd[item.name] = 0.0

    updated_count = 0
    for key, value in form.multi_items():
        if not key.startswith("upd_"):
            continue
        product_name = key[4:]
        try:
            new_upd = float(value)
        except (ValueError, TypeError):
            continue
        old = old_upd.get(product_name, 0.0)
        if abs(new_upd - old) < 0.001:
            continue
        batches = session.exec(
            select(StockItem).where(StockItem.name == product_name)
        ).all()
        for batch in batches:
            plan = plans_by_location.get(batch.storage_location)
            if plan:
                new_target = round(new_upd * plan.participants * plan.stock_duration_days)
                batch.target_unidoses_location = new_target
                batch.updated_at = datetime.now(UTC)
                session.add(batch)
                updated_count += 1
    session.commit()
    return RedirectResponse("/items/unidose-plan?m=saved", status_code=303)


@app.get("/api/items", response_model=list[ItemRead])
def api_list_items(
    request: Request,
    session: Session = Depends(get_session),
    user: User = Depends(_require_api_user),
    category: str | None = None,
    non_food_category: str | None = None,
):
    _ = request, user
    query = select(StockItem)
    if category == "food":
        query = query.where(
            or_(StockItem.item_category == "food", StockItem.item_category == None)  # noqa: E711
        )
    elif category is not None:
        query = query.where(StockItem.item_category == category)
    if non_food_category is not None:
        query = query.where(StockItem.non_food_category == non_food_category)
    items = session.exec(query).all()
    return [_to_read_model(item) for item in items]


@app.post("/api/items", response_model=ItemRead)
async def api_create_item(
    item_in: ItemCreate,
    session: Session = Depends(get_session),
    user: User = Depends(_require_api_user),
):
    item = StockItem(**item_in.model_dump(), user_id=user.id)
    session.add(item)
    session.commit()
    session.refresh(item)
    if item.quantity > 0:
        session.add(
            StockMovement(
                stock_item_id=item.id,
                user_id=user.id,
                delta=item.quantity,
                note="Initial stock quantity",
            )
        )
        session.commit()
    try:
        await calendar_service.schedule_renewal(user=user, item=item)
    except CalendarSyncError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    await _notify_telegram_operation_async(
        operation="api-item-created",
        actor=user.email,
        detail=(
            f"name={item.name} type={item.item_type} qty={item.quantity} "
            f"location={item.storage_location}"
        ),
    )
    return _to_read_model(item)


@app.get("/api/excel/stocks", response_model=list[ItemRead])
def api_excel_list_stocks(
    request: Request,
    session: Session = Depends(get_session),
    user: User = Depends(_require_excel_api_user),
):
    _ = request, user
    items = session.exec(
        select(StockItem)
        .order_by(StockItem.name, StockItem.storage_location, StockItem.expiry_date)
    ).all()
    return [_to_read_model(item) for item in items]


@app.put("/api/excel/stocks/{item_id}", response_model=ItemRead)
async def api_excel_update_stock(
    item_id: int,
    item_in: ItemCreate,
    session: Session = Depends(get_session),
    user: User = Depends(_require_excel_api_user),
):
    item = session.get(StockItem, item_id)
    if not item:
        raise HTTPException(status_code=404, detail="Stock item not found.")

    previous_quantity = item.quantity
    _apply_item_create_to_stock(item, item_in)
    session.add(item)
    quantity_delta = item.quantity - previous_quantity
    if quantity_delta:
        session.add(
            StockMovement(
                stock_item_id=item.id,
                user_id=user.id,
                delta=quantity_delta,
                note="Quantity updated via Excel API",
            )
        )
    session.commit()
    session.refresh(item)
    try:
        await calendar_service.schedule_renewal(user=user, item=item)
    except CalendarSyncError as exc:
        raise HTTPException(status_code=502, detail=str(exc)) from exc
    await _notify_telegram_operation_async(
        operation="excel-item-updated",
        actor=user.email,
        detail=(
            f"id={item.id} name={item.name} qty={item.quantity} "
            f"delta={quantity_delta} location={item.storage_location}"
        ),
    )
    return _to_read_model(item)


@app.post("/api/excel/stocks/upsert")
async def api_excel_upsert_stocks(
    payload: ExcelStockUpsertRequest,
    session: Session = Depends(get_session),
    user: User = Depends(_require_excel_api_user),
):
    created = 0
    updated = 0
    rows: list[ItemRead] = []

    for row in payload.rows:
        item_in = ItemCreate.model_validate(row.model_dump(exclude={"id"}))
        item = None
        if row.id is not None:
            item = session.get(StockItem, row.id)
            if not item:
                raise HTTPException(
                    status_code=404,
                    detail=f"Stock item not found for id={row.id}.",
                )
        else:
            item = _excel_match_existing_item(session, row=row)

        if item:
            previous_quantity = item.quantity
            _apply_item_create_to_stock(item, item_in)
            session.add(item)
            quantity_delta = item.quantity - previous_quantity
            if quantity_delta:
                session.add(
                    StockMovement(
                        stock_item_id=item.id,
                        user_id=user.id,
                        delta=quantity_delta,
                        note="Quantity updated via Excel API upsert",
                    )
                )
            updated += 1
        else:
            item = StockItem(**item_in.model_dump(), user_id=user.id)
            session.add(item)
            session.flush()
            if item.quantity > 0:
                session.add(
                    StockMovement(
                        stock_item_id=item.id,
                        user_id=user.id,
                        delta=item.quantity,
                        note="Initial stock quantity (Excel API)",
                    )
                )
            created += 1

        session.commit()
        session.refresh(item)
        try:
            await calendar_service.schedule_renewal(user=user, item=item)
        except CalendarSyncError as exc:
            raise HTTPException(status_code=502, detail=str(exc)) from exc
        await _notify_telegram_operation_async(
            operation="excel-upsert-row",
            actor=user.email,
            detail=(
                f"id={item.id} name={item.name} qty={item.quantity} "
                f"location={item.storage_location}"
            ),
        )
        rows.append(_to_read_model(item))

    return {"created": created, "updated": updated, "rows": rows}


@app.post("/api/barcode-lookup", response_model=BarcodeLookupResult)
async def api_barcode_lookup(
    payload: BarcodeLookupRequest,
    user: User = Depends(_require_api_user),
):
    _ = user
    return await barcode_service.lookup(payload.barcode, payload.item_type)


class _TablePDFRequest(BaseModel):
    title: str = "Export"
    filters: dict[str, str] = {}
    columns: list[str]
    rows: list[list[str]]


@app.post("/api/pdf/table")
async def api_pdf_table(
    payload: _TablePDFRequest,
    user: User = Depends(_require_api_user),
):
    """Generate and return an A4 PDF for a client-side filtered table."""
    _ = user
    import re as _re

    pdf_bytes = generate_table_pdf(
        title=payload.title,
        filters=payload.filters,
        columns=payload.columns,
        rows=payload.rows,
    )
    safe_name = _re.sub(r"[^a-z0-9_-]", "_", payload.title.lower())[:50] or "export"
    return StreamingResponse(
        io.BytesIO(pdf_bytes),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}.pdf"'},
    )


@app.post("/api/telegram/webhook")
async def telegram_webhook():
    raise HTTPException(
        status_code=410,
        detail=(
            "Telegram webhook integration moved out of FastAPI. "
            "Run scripts/telegram_copilot_bridge.py for Telegram <-> Copilot CLI chat."
        ),
    )


class BackupRestoreRequest(BaseModel):
    filename: str


def _get_db_file() -> Path | None:
    db_file = _sqlite_path_from_url(settings.database_url)
    if db_file is not None and not db_file.is_absolute():
        db_file = (Path.cwd() / db_file).resolve()
    return db_file


@app.post("/api/admin/backup")
async def api_admin_backup(current_user: User = Depends(_require_api_user)):
    if not current_user.is_admin:
        raise HTTPException(status_code=403, detail="Admin only")
    db_file = _get_db_file()
    backup_path = create_backup(db_file)
    if backup_path is None:
        raise HTTPException(status_code=404, detail="Database file not found")
    stat = backup_path.stat()
    return {
        "filename": backup_path.name,
        "size_bytes": stat.st_size,
        "created_at": datetime.now(UTC).isoformat(),
    }


@app.post("/api/admin/restore")
async def api_admin_restore(
    req: BackupRestoreRequest, current_user: User = Depends(_require_api_user)
):
    if not current_user.is_admin:
        raise HTTPException(403, "Admin only")
    db_file = _get_db_file()
    ok = restore_backup(db_file, req.filename)
    if not ok:
        raise HTTPException(404, detail=f"Backup not found: {req.filename}")
    return {
        "restored": True,
        "filename": req.filename,
        "message": "Database restored. Please reload the application.",
    }


@app.get("/api/admin/backups")
async def api_admin_list_backups(current_user: User = Depends(_require_api_user)):
    if not current_user.is_admin:
        raise HTTPException(403, "Admin only")
    db_file = _get_db_file()
    return {"backups": list_backups(db_file)}


@app.exception_handler(HTTPException)
async def http_exception_handler(_: Request, exc: HTTPException):
    return JSONResponse({"detail": exc.detail}, status_code=exc.status_code)
