from __future__ import annotations

import csv
import io
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

from app.schemas import ImportResult, ItemCreate

HEADER_MAP = {
    "barcode": "barcode",
    "batch_code": "batch_code",
    "batch": "batch_code",
    "lot": "batch_code",
    "name": "name",
    "item_name": "name",
    "type": "item_type",
    "item_type": "item_type",
    "storage_location": "storage_location",
    "location": "storage_location",
    "storage_bucket": "storage_bucket",
    "bucket": "storage_bucket",
    "expiry_date": "expiry_date",
    "expiry": "expiry_date",
    "quantity": "quantity",
    "qty": "quantity",
    "unidose_per_pack": "unidose_per_pack",
    "target_unidoses_location": "target_unidoses_location",
    "target_unidose_location": "target_unidoses_location",
    "temp_min_c": "temp_min_c",
    "temp_max_c": "temp_max_c",
    "humidity_min_pct": "humidity_min_pct",
    "humidity_max_pct": "humidity_max_pct",
    "renewal_date": "renewal_date",
}


def _canonical_header(value: str) -> str:
    return value.strip().lower().replace(" ", "_")


def _as_iso_date(value: object) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def _normalize_row(raw: dict[str, object]) -> dict[str, object]:
    normalized: dict[str, object] = {}
    for key, value in raw.items():
        canonical = HEADER_MAP.get(_canonical_header(key), None)
        if canonical is None:
            continue
        if canonical in {"expiry_date", "renewal_date"}:
            normalized[canonical] = _as_iso_date(value)
        elif canonical == "storage_bucket":
            normalized[canonical] = str(value or "")
        elif canonical == "quantity":
            normalized[canonical] = int(value) if value not in (None, "") else 0
        elif canonical == "unidose_per_pack":
            normalized[canonical] = int(value) if value not in (None, "") else 1
        elif canonical == "target_unidoses_location":
            normalized[canonical] = int(value) if value not in (None, "") else 0
        elif value == "":
            normalized[canonical] = None
        else:
            normalized[canonical] = value
    return normalized


def _parse_csv(file_bytes: bytes) -> list[dict[str, object]]:
    text_stream = io.StringIO(file_bytes.decode("utf-8-sig"))
    reader = csv.DictReader(text_stream)
    return [dict(row) for row in reader]


def _parse_excel(file_bytes: bytes) -> list[dict[str, object]]:
    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    parsed: list[dict[str, object]] = []
    for row in rows[1:]:
        parsed.append({headers[idx]: row[idx] for idx in range(len(headers)) if headers[idx]})
    return parsed


def parse_import_file(file_bytes: bytes, filename: str) -> tuple[list[ItemCreate], ImportResult]:
    suffix = Path(filename).suffix.lower()
    if suffix not in {".csv", ".xlsx"}:
        raise ValueError("Unsupported file format. Only CSV and XLSX are supported.")

    raw_rows = _parse_csv(file_bytes) if suffix == ".csv" else _parse_excel(file_bytes)
    items: list[ItemCreate] = []
    errors: list[str] = []

    for index, row in enumerate(raw_rows, start=2):
        payload = _normalize_row(row)
        try:
            item = ItemCreate.model_validate(payload)
            items.append(item)
        except Exception as exc:  # pydantic validation error details are propagated in message
            errors.append(f"Row {index}: {exc}")

    return items, ImportResult(imported=len(items), failed=len(errors), errors=errors)
